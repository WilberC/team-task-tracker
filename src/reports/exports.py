"""Excel exports for operational reports."""

from collections.abc import Iterable
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.reports.selectors import (
    AreaProductivity,
    DeadlineCompliance,
    StatusCount,
    WorkloadCount,
)
from src.tasks.models import Task

ACCENT = "E8732C"
INK = "0E1014"
STEEL_50 = "F1F2F5"
STEEL_200 = "D2D6DD"
STEEL_600 = "535A66"
STEEL_800 = "2C313A"
PAPER = "FAF8F4"
WHITE = "FFFFFF"


def build_reports_workbook(
    *,
    start_date: date,
    end_date: date,
    status_counts: Iterable[StatusCount],
    overdue_tasks: Iterable[Task],
    workload: Iterable[WorkloadCount],
    productivity: Iterable[AreaProductivity],
    deadline: DeadlineCompliance,
    due_date_tasks: Iterable[Task],
) -> bytes:
    status_counts = tuple(status_counts)
    overdue_tasks = tuple(overdue_tasks)
    workload = tuple(workload)
    productivity = tuple(productivity)
    due_date_tasks = tuple(due_date_tasks)

    workbook = Workbook()
    workbook.properties.creator = "Jawinsa"
    workbook.properties.title = "Reportes operativos"

    summary = workbook.active
    summary.title = "Resumen"
    _setup_sheet(summary)
    _write_title(
        summary,
        "Reportes operativos",
        f"Rango analizado: {start_date:%d/%m/%Y} - {end_date:%d/%m/%Y}",
    )
    _write_summary_cards(
        summary,
        [
            ("Tareas en rango", len(due_date_tasks)),
            ("Tareas vencidas", len(overdue_tasks)),
            ("A tiempo", f"{deadline.on_time_percent}%"),
            ("Tarde", f"{deadline.late_percent}%"),
        ],
    )
    _write_table(
        summary,
        start_row=8,
        title="Lectura rapida",
        headers=("Indicador", "Valor", "Detalle"),
        rows=(
            ("Cumplimiento a tiempo", deadline.on_time_percent, deadline.on_time_count),
            ("Cumplimiento tarde", deadline.late_percent, deadline.late_count),
            ("Responsables con carga abierta", len(workload), ""),
            ("Areas con productividad", len(productivity), ""),
        ),
    )

    _add_status_sheet(workbook, status_counts)
    _add_due_tasks_sheet(workbook, due_date_tasks, start_date, end_date)
    _add_overdue_sheet(workbook, overdue_tasks)
    _add_workload_sheet(workbook, workload)
    _add_productivity_sheet(workbook, productivity)
    _add_deadline_sheet(workbook, deadline)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _add_status_sheet(
    workbook: Workbook,
    status_counts: tuple[StatusCount, ...],
) -> None:
    sheet = workbook.create_sheet("Tareas por estado")
    _setup_sheet(sheet)
    _write_title(sheet, "Tareas por estado", "Agrupadas por fecha limite en el rango.")
    _write_table(
        sheet,
        start_row=5,
        title="Estados",
        headers=("Estado", "Cantidad"),
        rows=((item.label, item.count) for item in status_counts),
    )


def _add_due_tasks_sheet(
    workbook: Workbook,
    tasks: tuple[Task, ...],
    start_date: date,
    end_date: date,
) -> None:
    sheet = workbook.create_sheet("Detalle tareas")
    _setup_sheet(sheet)
    _write_title(
        sheet,
        "Detalle de tareas",
        f"Tareas con fecha limite entre {start_date:%d/%m/%Y} y {end_date:%d/%m/%Y}.",
    )
    _write_table(
        sheet,
        start_row=5,
        title="Tareas filtradas",
        headers=(
            "Tarea",
            "Tipo",
            "Orden",
            "Vehiculo",
            "Cliente",
            "Area",
            "Responsable",
            "Tipo responsable",
            "Prioridad",
            "Estado",
            "Fecha inicio",
            "Fecha limite",
            "Fecha finalizacion",
        ),
        rows=(_task_row(task) for task in tasks),
    )


def _add_overdue_sheet(workbook: Workbook, tasks: tuple[Task, ...]) -> None:
    sheet = workbook.create_sheet("Tareas vencidas")
    _setup_sheet(sheet)
    _write_title(sheet, "Tareas vencidas", "Trabajos abiertos fuera de fecha.")
    _write_table(
        sheet,
        start_row=5,
        title="Vencimientos",
        headers=(
            "Tarea",
            "Vehiculo",
            "Area",
            "Responsable",
            "Prioridad",
            "Fecha limite",
        ),
        rows=(
            (
                task.title,
                _vehicle_plate(task),
                task.area.name,
                _assignee(task)[1],
                task.get_priority_display(),
                task.due_date,
            )
            for task in tasks
        ),
    )


def _add_workload_sheet(
    workbook: Workbook,
    workload: tuple[WorkloadCount, ...],
) -> None:
    sheet = workbook.create_sheet("Carga")
    _setup_sheet(sheet)
    _write_title(
        sheet,
        "Carga por responsable",
        "Tareas abiertas por mecanico o equipo.",
    )
    _write_table(
        sheet,
        start_row=5,
        title="Carga activa",
        headers=("Responsable", "Tipo", "Tareas abiertas"),
        rows=((item.assignee, item.kind, item.count) for item in workload),
    )


def _add_productivity_sheet(
    workbook: Workbook,
    productivity: tuple[AreaProductivity, ...],
) -> None:
    sheet = workbook.create_sheet("Productividad")
    _setup_sheet(sheet)
    _write_title(sheet, "Productividad por area", "Tareas completadas en el rango.")
    _write_table(
        sheet,
        start_row=5,
        title="Productividad",
        headers=("Area", "Tareas completadas"),
        rows=((item.area, item.completed) for item in productivity),
    )


def _add_deadline_sheet(workbook: Workbook, deadline: DeadlineCompliance) -> None:
    sheet = workbook.create_sheet("Cumplimiento")
    _setup_sheet(sheet)
    _write_title(
        sheet,
        "Cumplimiento de fechas",
        "Completadas a tiempo frente a tarde.",
    )
    _write_table(
        sheet,
        start_row=5,
        title="Cumplimiento",
        headers=("Resultado", "Cantidad", "Porcentaje"),
        rows=(
            ("A tiempo", deadline.on_time_count, deadline.on_time_percent / 100),
            ("Tarde", deadline.late_count, deadline.late_percent / 100),
        ),
    )
    for cell in sheet["C"][5:]:
        cell.number_format = "0%"


def _setup_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 18


def _write_title(sheet, title: str, subtitle: str) -> None:
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Archivo", size=20, bold=True, color=INK)
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="Manrope", size=11, bold=True, color=STEEL_600)
    sheet["A3"] = "Jawinsa"
    sheet["A3"].font = Font(name="Manrope", size=10, bold=True, color=ACCENT)


def _write_summary_cards(sheet, cards: list[tuple[str, object]]) -> None:
    for index, (label, value) in enumerate(cards, start=1):
        column = (index - 1) * 2 + 1
        cell = sheet.cell(row=5, column=column, value=label)
        value_cell = sheet.cell(row=6, column=column, value=value)
        sheet.merge_cells(
            start_row=5,
            start_column=column,
            end_row=5,
            end_column=column + 1,
        )
        sheet.merge_cells(
            start_row=6,
            start_column=column,
            end_row=6,
            end_column=column + 1,
        )
        for row in (5, 6):
            for offset in (0, 1):
                current = sheet.cell(row=row, column=column + offset)
                current.fill = PatternFill("solid", fgColor=PAPER)
                current.border = _border()
                current.alignment = Alignment(horizontal="center")
        cell.font = Font(name="Manrope", size=10, bold=True, color=STEEL_600)
        value_cell.font = Font(name="Archivo", size=22, bold=True, color=INK)


def _write_table(
    sheet,
    *,
    start_row: int,
    title: str,
    headers: tuple[str, ...],
    rows: Iterable[tuple[object, ...]],
) -> None:
    sheet.cell(row=start_row - 1, column=1, value=title)
    sheet.cell(row=start_row - 1, column=1).font = Font(
        name="Archivo",
        size=14,
        bold=True,
        color=INK,
    )
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column_index, value=header)
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.font = Font(name="Manrope", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center")
        cell.border = _border(color=INK)

    body_rows = list(rows) or [
        tuple("Sin datos" if index == 0 else "" for index in range(len(headers)))
    ]
    for row_index, row in enumerate(body_rows, start=start_row + 1):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            fill_color = WHITE if row_index % 2 else STEEL_50
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(name="Manrope", size=10, color=STEEL_800)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _border()
            if isinstance(value, date):
                cell.number_format = "dd/mm/yyyy"

    end_row = start_row + len(body_rows)
    end_column = len(headers)
    reference = f"A{start_row}:{get_column_letter(end_column)}{end_row}"
    sheet.auto_filter.ref = reference
    _fit_columns(sheet, end_column, end_row)


def _fit_columns(sheet, end_column: int, end_row: int) -> None:
    for column_index in range(1, end_column + 1):
        letter = get_column_letter(column_index)
        values = [
            sheet.cell(row=row, column=column_index).value
            for row in range(1, end_row + 1)
        ]
        width = max(len(str(value)) if value is not None else 0 for value in values)
        sheet.column_dimensions[letter].width = min(max(width + 3, 14), 34)


def _border(color: str = STEEL_200) -> Border:
    side = Side(style="thin", color=color)
    return Border(top=side, right=side, bottom=side, left=side)


def _task_row(task: Task) -> tuple[object, ...]:
    assignee_kind, assignee = _assignee(task)
    return (
        task.title,
        "Subtarea" if task.is_subtask else "Tarea principal",
        _job_order_code(task),
        _vehicle_plate(task),
        _client_name(task),
        task.area.name,
        assignee,
        assignee_kind,
        task.get_priority_display(),
        task.get_status_display(),
        task.start_date,
        task.due_date,
        task.completion_date,
    )


def _assignee(task: Task) -> tuple[str, str]:
    if task.assigned_employee_id:
        return ("Empleado", task.assigned_employee.full_name)
    if task.assigned_team_id:
        return ("Equipo", task.assigned_team.name)
    return ("Sin asignar", "")


def _job_order_code(task: Task) -> str:
    if task.job_order_id:
        return f"OT-{task.job_order_id}"
    if task.parent_task_id and task.parent_task.job_order_id:
        return f"OT-{task.parent_task.job_order_id}"
    return ""


def _vehicle_plate(task: Task) -> str:
    if task.job_order_id:
        return task.job_order.vehicle.plate
    if task.parent_task_id and task.parent_task.job_order_id:
        return task.parent_task.job_order.vehicle.plate
    return ""


def _client_name(task: Task) -> str:
    if task.job_order_id:
        return task.job_order.service_order.client.full_name
    if task.parent_task_id and task.parent_task.job_order_id:
        return task.parent_task.job_order.service_order.client.full_name
    return ""
